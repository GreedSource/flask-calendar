from datetime import datetime, timedelta, date
from random import randrange, choice
# Idioma "es-ES" (código para el español de España)

import xlrd, json
from __class.weekday import process_data
import openpyxl

class xlsx_reader(object):

    def json_output(self, data):
        date = datetime.now()
        name = 'calendario'+ date.strftime('%Y%m%d-%H%M%S')
        with open(f'json/{name}.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return True
        
    def read(self, file):
        all_data = []
        excel = xlrd.open_workbook(file)
        sheet_0 = excel.sheet_by_index(0) # Open the first tab
        prev_row = [None for i in range(sheet_0.ncols)]
        first = True
        for row_index in range(sheet_0.nrows):
            if first == False:
                row = []
                for col_index in range(sheet_0.ncols):
                    value = sheet_0.cell(rowx=row_index,colx=col_index).value
                    if len(value) == 0:
                        value = prev_row[col_index]
                    row.append(value)
                prev_row = row
                all_data.append(row)
            else:
                first = False
            ######
        data = []
        for row in all_data:
                #data.append(row)
            tmp = []
            for child in row:
                if child and '\n' in child:
                    child = child.split('\n')
                tmp.append(child)
            data.append(tmp)
        data = self.parse_to_json(data)
        return data

    def parse_to_json(self, data):
        array = {}
        array['lunes'] = []
        array['martes'] = []
        array['miercoles'] = []
        array['jueves'] = []
        array['viernes'] = []
        for hour in data:
            if type(hour[1]) != str:
                existe = False
                x = 0
                for major in array['lunes']:
                    #print(major)
                    if major[1] == hour[1][1]:
                        x = array['lunes'].index(major)
                        existe = True
                        #print(f'{major[1]} - {hour[1][1]}')
                        break
                if existe == True and major[1] in hour[1]:
                    y = len(hour[1])
                    array['lunes'][x][y] += 1
                else:
                    tmp = hour[1]
                    tmp.append(1)
                    tmp.append(hour[0])
                    array['lunes'].append(tmp)
            
            if type(hour[2]) != str:
                existe = False
                x = 0
                for major in array['martes']:
                    #print(major)
                    if major[1] == hour[2][1]:
                        x = array['martes'].index(major)
                        existe = True
                        #print(f'{major[1]} - {hour[1][1]}')
                        break
                if existe == True and major[1] in hour[2]:
                    y = len(hour[2])
                    array['martes'][x][y] += 1
                else:
                    tmp = hour[2]
                    tmp.append(1)
                    tmp.append(hour[0])
                    array['martes'].append(tmp)

            if type(hour[3]) != str:
                existe = False
                x = 0
                for major in array['miercoles']:
                    #print(major)
                    if major[1] == hour[3][1]:
                        x = array['miercoles'].index(major)
                        existe = True
                        #print(f'{major[1]} - {hour[3][1]}')
                        break
                if existe == True and major[1] in hour[3]:
                    y = len(hour[3])
                    array['miercoles'][x][y] += 1
                else:
                    tmp = hour[3]
                    tmp.append(1)
                    tmp.append(hour[0])
                    array['miercoles'].append(tmp)
            
            if type(hour[4]) != str:
                existe = False
                x = 0
                for major in array['jueves']:
                    #print(major)
                    if major[1] == hour[4][1]:
                        x = array['jueves'].index(major)
                        existe = True
                        #print(f'{major[1]} - {hour[1][1]}')
                        break
                if existe == True and major[1] in hour[4]:
                    y = len(hour[4])
                    array['jueves'][x][y] += 1
                else:
                    tmp = hour[4]
                    tmp.append(1)
                    tmp.append(hour[0])
                    array['jueves'].append(tmp)
            if type(hour[5]) != str:
                existe = False
                x = 0
                for major in array['viernes']:
                    #print(major)
                    if major[1] == hour[5][1]:
                        x = array['viernes'].index(major)
                        existe = True
                        #print(f'{major[1]} - {hour[1][1]}')
                        break
                if existe == True and major[1] in hour[5]:
                    y = len(hour[5])
                    array['viernes'][x][y] += 1
                else:
                    tmp = hour[5]
                    tmp.append(1)
                    tmp.append(hour[0])
                    array['viernes'].append(tmp)
        
        array = self.obtain_majors(array)
        self.json_output(array)
        return array

    def obtain_majors(self, data):
        array = []
        for record in data:
            for row in data[record]:
                    if not (row[1] in array):
                        if row[1] != 'Tutorías':
                            hour = len(row) - 2
                            hours = hour + 1
                            tmp = [row[0], row[1], row[hours], row[hour]]
                            array.append(tmp)
                            #array.append(tmp)
        #print(array)
        
        output = {}
        for record in data:
            for major in array:
                    if not major[1] in output:
                        for row in data[record]:      
                                if row[1] == major[1]:
                                    hours = row[len(row)-2]
                                    if hours > 1:
                                        #print(row)
                                        if len(row[1]) > 0:
                                            x = len(row) - 1
                                            tmp = [record, row[0], row[x]]
                                            output[major[1]] = [tmp]
                                        else:
                                            x = len(row) - 1
                                            tmp = [record, row[0], row[x]]
                                            output[major[1]] = [tmp]
                                        #print(record)
                    else:
                        for row in data[record]:      
                                if row[1] == major[1]:
                                    hours = row[len(row)-2]
                                    if hours > 1:
                                        if len(row[1]) > 0:
                                            x = len(row) - 1
                                            tmp = [record, row[0], row[x]]
                                            output[major[1]].append(tmp)
                                        else:
                                            x = len(row) - 1
                                            tmp = [record, row[0], row[x]]
                                            output[major[1]].append(tmp)
        exit = {}
        for record in output:
            exit[record] = []
            for row in output[record]:
                if any(list == row for list in exit[record]) == False:
                    exit[record].append(row)
        return exit

class xlsx_writer(object):
    def __init__(self, carrera, grado, grupo):
        self.carrera = carrera
        self.grado = grado
        self.grupo = grupo
        self.carrera = carrera
        import shutil
        fuente = "resources/template.xlsx"
        date = datetime.now().strftime("%Y-%m-%d-%H%M%S")
        name = f'calendario-{carrera}-{grado}{grupo}-{date}'
        name = f'output/{name}.xlsx'
        self.name = name
        shutil.copyfile(fuente, name)

    def date_assignament(self, majors, corte, habiles):
        array = {}
        for major in majors:
            
            #print(major)

            wd = process_data()
            
            #print(f'{_major} - {day}')

            for row in corte:
                _major = choice(majors[major])
                day = wd.date_assignament(_major[0])
                
                #print(row)
                if row.weekday() == day:
                    if not major in array:
                        tmp = [row, _major[1], _major[2]]
                        array[major] = [tmp]
                    else:
                        #print('append')
                        tmp = [row, _major[1], _major[2]]
                        array[major].append(tmp)
                else:
                    
                    date = self.__recursive_date_validation(row, day, habiles)
                    if not major in array:
                        tmp = [date, _major[1], _major[2]]
                        array[major] = [tmp]
                    else:
                        tmp = [date, _major[1], _major[2]]
                        array[major].append(tmp)

        #self.json_output(array)
        
        return array

    def __recursive_date_validation(self, date, wd, habiles):
        date = date - timedelta(days=1)
        if date.weekday() == wd:
            if date in habiles:
                return date
            else:
                return self.__recursive_date_validation(date, wd, habiles)
        else:
            return self.__recursive_date_validation(date, wd, habiles)

    def __re_asign_date(self, majors, dates, actual, corte):
        wd = process_data()
        __date = actual.weekday()
        print(corte)

        details = None
        evaluate = None
        
        if len(majors) > 1:
            for major in majors:
                tmp = wd.date_assignament(major[0])
                if __date != tmp:
                    evaluate = tmp
                    details = [major[1], major[2]]
                    break
        else:
            return None
        final_date = None
        for dt in dates:
            if dt > actual and dt < corte and dt.weekday() == evaluate:
                final_date = dt
                break
        
        if final_date != None:
            return [final_date, details[0], details[1]]
        else:
            return None

        

    def write(self, corte, majors, habiles, entregas, ciclo):
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        wb = openpyxl.load_workbook(self.name)
        ws = wb.worksheets[0]
        #img = openpyxl.drawing.image.Image('resources/logo.png')
        #img.anchor(ws.cell('A1'))
        #img.anchor = 'A1'
        #ws.add_image(img)
        ws['B2'] = f'{self.grado}-{self.grupo}'
        ws['B3'] = f'{self.carrera} - {ciclo} - {date.today().strftime("%Y")}'
        index = 1
        for item in entregas:
            #item = item.strftime('%d-%m-%Y')
            cell = f'{columns[index]}8'
            ws[cell] = item
            index += 1
        index = 9
        for major in majors:
            cell = f'A{index}'
            ws[cell] = major
            index += 1
        
        
        fechas = self.date_assignament(majors, corte, habiles)

        row = 9
        #print(majors)
        for major in fechas:
            column = 1
            tmp = None
            index = 0
            for fecha in fechas[major]:
                pendiente = 0
                if fecha[0] == tmp:
                    new_date = self.__re_asign_date(majors[major], habiles, fecha[0], corte[index])
                    
                    if new_date != None:
                        dt = new_date[0].strftime('%A %d de %B')
                        lab = new_date[1]
                        hora = new_date[2].split('-')[0]
                    else:
                        dt = 'Pendiente de\nvalidar con maestro'
                        pendiente = 1
                elif tmp != None and fecha[0] < tmp:
                    new_date = self.__re_asign_date(majors[major], habiles, fecha[0], corte[index])
                    
                    if new_date != None:
                        dt = new_date[0].strftime('%A %d de %B')
                        lab = new_date[1]
                        hora = new_date[2].split('-')[0]
                    else:
                        dt = 'Pendiente de\nvalidar con maestro'
                        pendiente = 1
                else:
                    dt = fecha[0].strftime('%A %d de %B')
                    lab = fecha[1]
                    hora = fecha[2].split('-')[0]
                
                cell = f'{columns[column]}{row}'
                
                if pendiente == 0:
                    value = f'{dt}\n{hora}\n{lab}'
                else:
                    value = f'{dt}'
                
                ws[cell] = value
                column += 1
                tmp = fecha[0]
                index += 1
            row += 1
        
        wb.save(self.name)